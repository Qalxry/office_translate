"""CLI（任务制）全流程测试：init / list / extract / apply。"""

import openpyxl
import pytest

from office_translate.cli import main

CONFIG = "work_dir: work\noutput_dir: output\nsep: '\\n'\n"


@pytest.fixture
def project(tmp_path):
    """临时项目：config.yaml + 一个 sample.xlsx 输入。"""
    (tmp_path / "config.yaml").write_text(CONFIG, encoding="utf-8")
    in_dir = tmp_path / "input"
    in_dir.mkdir()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Hello"
    ws["B1"] = "World"
    wb.save(in_dir / "sample.xlsx")
    return tmp_path


def test_full_flow(project):
    root = project
    cfg = str(root / "config.yaml")
    sample = str(root / "input" / "sample.xlsx")

    # init
    assert main(["-c", cfg, "init", "job1", "-i", sample]) == 0
    job_dir = root / "work" / "job1"
    assert (job_dir / "sample.xlsx").is_file()
    assert (job_dir / "job.yaml").is_file()

    # 重复 init 报错
    assert main(["-c", cfg, "init", "job1", "-i", sample]) == 1

    # init 不存在的输入报错
    assert main(["-c", cfg, "init", "job2", "-i", str(root / "nope.xlsx")]) == 1

    # list 能看到任务
    assert main(["-c", cfg, "list"]) == 0

    # extract
    assert main(["-c", cfg, "extract", "job1"]) == 0
    assert (job_dir / "source.txt").is_file()
    assert (job_dir / "map.json").is_file()

    # 未翻译时 apply 报错
    assert main(["-c", cfg, "apply", "job1"]) == 1

    # 模拟人工翻译
    src = (job_dir / "source.txt").read_text(encoding="utf-8")
    (job_dir / "translated.txt").write_text(
        src.replace("Hello", "你好").replace("World", "世界"), encoding="utf-8"
    )

    # apply
    assert main(["-c", cfg, "apply", "job1"]) == 0
    out_dir = job_dir / "output"
    assert (out_dir / "job1_translated.xlsx").is_file()
    assert (out_dir / "job1_bilingual.xlsx").is_file()

    # 校验回填结果
    wb = openpyxl.load_workbook(out_dir / "job1_translated.xlsx")
    assert wb.active["A1"].value == "你好"
    wb2 = openpyxl.load_workbook(out_dir / "job1_bilingual.xlsx")
    assert wb2.active["A1"].value == "Hello\n你好"


def test_extract_missing_job(project):
    assert main(["-c", str(project / "config.yaml"), "extract", "ghost"]) == 1


def test_init_auto_job_name(project):
    import re
    root = project
    cfg = str(root / "config.yaml")
    sample = str(root / "input" / "sample.xlsx")

    # 不带 job：自动按时间戳命名
    assert main(["-c", cfg, "init", "-i", sample]) == 0
    dirs = sorted(d.name for d in (root / "work").iterdir() if d.is_dir())
    assert len(dirs) == 1
    assert re.fullmatch(r"\d{8}_\d{6}", dirs[0]), f"任务名应为时间戳: {dirs[0]}"
    assert (root / "work" / dirs[0] / "job.yaml").is_file()


def test_init_normalizes_nbsp_filename(project):
    # 文件名含不可断行空格（U+00A0）时，复制进任务目录应替换为普通空格
    root = project
    weird = root / "input" / "a b.xlsx"
    wb = openpyxl.Workbook()
    wb.save(weird)

    cfg = str(root / "config.yaml")
    assert main(["-c", cfg, "init", "job_nbsp", "-i", str(weird)]) == 0
    job_dir = root / "work" / "job_nbsp"
    assert (job_dir / "a b.xlsx").is_file()
    assert not (job_dir / "a b.xlsx").exists()


def test_config_missing_uses_defaults(tmp_path):
    # 无 config.yaml 也能跑（用内置默认值，work_dir 相对当前目录）
    assert main(["-c", str(tmp_path / "nonexistent.yaml"), "list"]) == 0


def test_auto_translate_google(project, monkeypatch):
    # mock GoogleProvider，避免真请求
    from office_translate.ai import provider as provider_mod

    class FakeGoogle:
        name = "google"

        def __init__(self, mirrors):
            pass

        def translate_batch(self, texts, source, target, concurrency=1):
            return [f"译:{t}" for t in texts]

    monkeypatch.setattr(provider_mod, "GoogleProvider", FakeGoogle)

    root = project
    cfg = str(root / "config.yaml")
    sample = str(root / "input" / "sample.xlsx")
    assert main(["-c", cfg, "init", "jobauto", "-i", sample]) == 0

    # 一键翻译
    assert main(["-c", cfg, "auto", "jobauto"]) == 0

    job_dir = root / "work" / "jobauto"
    assert (job_dir / "source.txt").is_file()
    assert (job_dir / "translated.txt").is_file()
    assert (job_dir / "output" / "jobauto_translated.xlsx").is_file()

    # 校验译文
    wb = openpyxl.load_workbook(job_dir / "output" / "jobauto_translated.xlsx")
    assert wb.active["A1"].value == "译:Hello"


def test_init_auto_job_suffix_on_collision(project, monkeypatch):
    # 时间戳重名时自动追加 _1 后缀
    import re
    from office_translate import cli as cli_mod

    root = project
    cfg = str(root / "config.yaml")
    sample = str(root / "input" / "sample.xlsx")

    class FakeDateTime:
        @staticmethod
        def now():
            return type("F", (), {"strftime": staticmethod(lambda fmt: "20260810_120000")})()
    monkeypatch.setattr(cli_mod, "datetime", FakeDateTime)
    assert main(["-c", cfg, "init", "-i", sample]) == 0
    assert main(["-c", cfg, "init", "-i", sample]) == 0
    names = sorted(d.name for d in (root / "work").iterdir() if d.is_dir())
    assert names == ["20260810_120000", "20260810_120000_1"], names


def test_init_xls_non_windows_prompts_manual(project, monkeypatch, capsys):
    # 非 Windows + .xls 输入：提示手动转换并返回 1，不建任务
    from office_translate import cli as cli_mod

    root = project
    cfg = str(root / "config.yaml")
    fake_xls = root / "input" / "questionnaire.xls"
    fake_xls.write_bytes(b"fake xls content")

    monkeypatch.setattr(cli_mod.win_convert, "_is_windows", lambda: False)
    assert main(["-c", cfg, "init", "-i", str(fake_xls)]) == 1
    out = capsys.readouterr().out
    assert "手动转换" in out and "另存为" in out
    assert not (root / "work").exists()  # 未创建任务
