"""File-level XLSX preflight, rich-text and fidelity regressions."""

from __future__ import annotations

import zipfile

import openpyxl
import pytest
from openpyxl.chart import BarChart, Reference
from openpyxl.cell.rich_text import CellRichText, InlineFont, TextBlock
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation

from office_translate.artifacts import TranslationArtifact
from office_translate.formats.xlsx import applier
from office_translate.formats.xlsx.applier import apply_artifacts
from office_translate.formats.xlsx.extractor import (
    extract_artifact,
    preflight_workbook,
)


def _rich_workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "翻译"
    ws["A1"] = CellRichText(
        [
            "Hello ",
            TextBlock(InlineFont(b=True, color="FFFF0000"), "world"),
        ]
    )
    wb.save(path)


def test_preflight_distinguishes_unsupported_and_corrupt_inputs(tmp_path):
    old_format = tmp_path / "book.xls"
    old_format.write_bytes(b"not an xls")
    unsupported = preflight_workbook(old_format)
    assert unsupported.kind == "unsupported"
    assert not unsupported.ok
    assert unsupported.diagnostics[0].code == "xlsx.unsupported_extension"

    fake = tmp_path / "fake.xlsx"
    fake.write_bytes(b"not an OOXML workbook")
    corrupt = preflight_workbook(fake)
    assert corrupt.kind == "corrupt"
    assert not corrupt.ok
    assert corrupt.diagnostics[0].code == "xlsx.invalid_zip"


def test_preflight_reports_rich_text_protection_and_ooxml_objects(tmp_path):
    path = tmp_path / "diagnostics.xlsx"
    _rich_workbook(path)
    wb = openpyxl.load_workbook(path, rich_text=True)
    wb.active.protection.sheet = True
    wb.save(path)

    # Add a harmless OOXML object part to exercise the structured warning
    # without requiring Excel-specific drawing construction.
    with zipfile.ZipFile(path, "a") as archive:
        archive.writestr("xl/embeddings/unverified.bin", b"object")

    result = preflight_workbook(path)
    assert result.ok
    codes = {diagnostic.code for diagnostic in result.diagnostics}
    assert "xlsx.rich_text_runs" in codes
    assert "xlsx.protection" in codes
    assert "xlsx.unverified_ooxml_objects" in codes
    assert all(diagnostic.to_dict()["action"] for diagnostic in result.diagnostics)


def test_multirun_rich_text_defaults_to_flatten_and_can_preserve_original(tmp_path):
    path = tmp_path / "rich.xlsx"
    _rich_workbook(path)
    source = extract_artifact(path)
    translation = TranslationArtifact.create(
        source,
        [{"id": item.id, "translation": "你好世界"} for item in source.items],
    )
    translated = tmp_path / "translated.xlsx"
    bilingual = tmp_path / "bilingual.xlsx"

    apply_artifacts(path, source, translation, translated, bilingual)
    assert openpyxl.load_workbook(translated).active["A1"].value == "你好世界"

    apply_artifacts(
        path,
        source,
        translation,
        translated,
        bilingual,
        rich_text_policy="preserve_original",
    )
    preserved = openpyxl.load_workbook(translated, rich_text=True).active["A1"]
    assert isinstance(preserved.value, CellRichText)
    assert str(preserved.value) == "Hello world"

    flattened = tmp_path / "flattened.xlsx"
    flattened_bilingual = tmp_path / "flattened-bilingual.xlsx"
    apply_artifacts(
        path,
        source,
        translation,
        flattened,
        flattened_bilingual,
        rich_text_policy="flatten",
    )
    assert openpyxl.load_workbook(flattened).active["A1"].value == "你好世界"


def test_fidelity_roundtrip_keeps_common_xlsx_features(tmp_path):
    path = tmp_path / "features.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws["A1"] = "Title"
    ws["A1"].font = openpyxl.styles.Font(bold=True, color="FFFF0000")
    ws["A2"] = "Formula label"
    ws["B2"] = "=1+1"
    ws["A3"] = "Merged"
    ws.merge_cells("A3:B3")
    ws["A4"] = "Pick"
    ws["A4"].hyperlink = "https://example.com"
    ws["A4"].comment = Comment("review", "tester")
    ws.freeze_panes = "B2"
    validation = DataValidation(type="list", formula1='"A,B"')
    ws.add_data_validation(validation)
    validation.add(ws["B4"])
    ws["C1"] = 1
    ws["C2"] = 2
    chart = BarChart()
    chart.add_data(Reference(ws, min_col=3, min_row=1, max_row=2))
    ws.add_chart(chart, "E2")
    wb.save(path)

    source = extract_artifact(path)
    translation = TranslationArtifact.create(
        source,
        [{"id": item.id, "translation": f"译-{item.text}"} for item in source.items],
    )
    output = tmp_path / "features-out.xlsx"
    bilingual = tmp_path / "features-bi.xlsx"
    apply_artifacts(path, source, translation, output, bilingual)

    result = openpyxl.load_workbook(output)
    result_ws = result["Sheet"]
    assert result_ws["A1"].font.bold is True
    assert result_ws["A1"].font.color.rgb in ("FFFF0000", "00FFFF0000")
    assert result_ws["B2"].value == "=1+1"
    assert any(str(rng) == "A3:B3" for rng in result_ws.merged_cells.ranges)
    assert result_ws["A4"].hyperlink.target == "https://example.com"
    assert result_ws["A4"].comment.text == "review"
    assert result_ws.freeze_panes == "B2"
    assert len(result_ws.data_validations.dataValidation) == 1
    assert len(result_ws._charts) == 1


def test_atomic_output_failure_keeps_existing_outputs(tmp_path, monkeypatch):
    path = tmp_path / "atomic.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "Hello"
    wb.save(path)
    source = extract_artifact(path)
    translation = TranslationArtifact.create(source, [{"id": 0, "translation": "你好"}])
    translated = tmp_path / "atomic-translated.xlsx"
    bilingual = tmp_path / "atomic-bilingual.xlsx"
    translated.write_bytes(b"old translated")
    bilingual.write_bytes(b"old bilingual")

    original_build = applier._build_candidate

    def fail_bilingual(*args, **kwargs):
        if kwargs.get("mode") == "bilingual" or (len(args) >= 5 and args[4] == "bilingual"):
            raise OSError("simulated serializer failure")
        return original_build(*args, **kwargs)

    monkeypatch.setattr(applier, "_build_candidate", fail_bilingual)
    with pytest.raises(OSError, match="simulated"):
        apply_artifacts(path, source, translation, translated, bilingual)
    assert translated.read_bytes() == b"old translated"
    assert bilingual.read_bytes() == b"old bilingual"
    assert not list(tmp_path.glob(".*.xlsx"))
