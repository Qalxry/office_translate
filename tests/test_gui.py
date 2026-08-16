"""GUI 后端（FastAPI）接口测试：任务/提取/翻译/审核/apply/术语库。"""

import os
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest


def _extract(client, job: str) -> dict:
    response = client.post(f"/api/jobs/{job}/extract")
    assert response.status_code == 200, response.text
    return response.json()


def _save_translations(
    client,
    job: str,
    source_revision: str,
    translations: list[str],
) -> dict:
    results = [
        {
            "id": item_id,
            "translation": translation,
            "status": "succeeded",
            "error": None,
        }
        for item_id, translation in enumerate(translations)
    ]
    ids = list(range(len(results)))
    response = client.post(
        f"/api/jobs/{job}/ai_output",
        json={
            "source_revision": source_revision,
            "results": results,
            "summary": {
                "status": "succeeded",
                "total": len(results),
                "succeeded": len(results),
                "failed": 0,
                "cancelled": 0,
                "succeeded_ids": ids,
                "failed_ids": [],
                "cancelled_ids": [],
            },
        },
    )
    assert response.status_code == 200, response.text
    return response.json()


def _apply_current(
    client,
    job: str,
    source_revision: str,
    translation_revision: str,
    **extra,
):
    return client.post(
        f"/api/jobs/{job}/apply",
        json={
            "source_revision": source_revision,
            "translation_revision": translation_revision,
            **extra,
        },
    )


def test_health(client):
    r = client.get("/api/jobs")
    assert r.status_code == 200
    assert r.json() == []


def test_upload_file(client, tmp_path):
    # 浏览器原生文件选择器 → 上传 → 保存到 input/
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "上传测试"
    payload = BytesIO()
    workbook.save(payload)
    r = client.post(
        "/api/upload",
        files={"file": ("问卷.xlsx", payload.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["filename"] == "问卷.xlsx"
    assert data["path"].endswith("问卷.xlsx")
    assert os.path.isfile(data["path"])
    assert open(data["path"], "rb").read()[:2] == b"PK"


def test_upload_rejects_fake_xlsx_before_publishing(client):
    response = client.post(
        "/api/upload",
        files={"file": ("fake.xlsx", b"not-an-ooxml-zip", "application/octet-stream")},
    )
    assert response.status_code == 400
    assert "有效的 OOXML ZIP" in response.json()["detail"]


def test_gui_rejects_legacy_xls_with_save_as_guidance(client, tmp_path):
    uploaded = client.post(
        "/api/upload",
        files={"file": ("legacy.xls", b"not-an-xls", "application/octet-stream")},
    )
    assert uploaded.status_code == 400
    assert "另存为" in uploaded.json()["detail"]

    created = client.post(
        "/api/jobs",
        json={"job": "legacy", "input": str(tmp_path / "legacy.xls")},
    )
    assert created.status_code == 400
    assert "另存为" in created.json()["detail"]


def test_create_job_auto_name(client, tmp_path):
    # 不传 job → 自动按时间戳命名
    r = client.post("/api/jobs", json={"job": None, "input": str(tmp_path / "input" / "sample.xlsx")})
    assert r.status_code == 200, r.text
    info = r.json()
    import re
    assert re.fullmatch(r"\d{8}_\d{6}", info["job"]), info["job"]
    assert os.path.isdir(tmp_path / "work" / info["job"])


def test_create_and_extract(client, tmp_path):
    # 新建任务
    r = client.post("/api/jobs", json={"job": "job1", "input": str(tmp_path / "input" / "sample.xlsx")})
    assert r.status_code == 200, r.text
    assert (tmp_path / "work" / "job1" / "job.yaml").is_file()

    # 提取
    info = _extract(client, "job1")
    assert info["unique_texts"] == 2

    # 读 source
    r = client.get("/api/jobs/job1/source")
    assert r.status_code == 200
    assert r.json()["texts"] == ["Hello", "World"]

    # 保存译文
    saved = _save_translations(
        client,
        "job1",
        info["source_revision"],
        ["你好", "世界"],
    )

    # 读取译文（translated_file）
    r = client.get("/api/jobs/job1/translated_file")
    assert r.status_code == 200
    assert [item["translation"] for item in r.json()["items"]] == ["你好", "世界"]

    # apply
    r = _apply_current(
        client,
        "job1",
        info["source_revision"],
        saved["translation_revision"],
        sep="\\n",
    )
    assert r.status_code == 200, r.text
    out = r.json()
    assert out["translated_output"].endswith(".translated.xlsx")
    wb = openpyxl.load_workbook(out["translated_output"])
    assert wb.active["A1"].value == "你好"


@pytest.mark.p0_regression
def test_ai_output_requires_summary_and_partial_summary_blocks_export(
    client,
    app_workspace,
):
    job = "summary_gate"
    assert client.post(
        "/api/jobs",
        json={"job": job, "input": str(app_workspace.sample_xlsx)},
    ).status_code == 200
    extracted = _extract(client, job)
    results = [
        {"id": 0, "translation": "你好", "status": "succeeded"},
        {
            "id": 1,
            "translation": "World",
            "status": "failed",
            "error": "truncated",
        },
    ]
    missing_summary = client.post(
        f"/api/jobs/{job}/ai_output",
        json={"source_revision": extracted["source_revision"], "results": results},
    )
    assert missing_summary.status_code == 400

    saved = client.post(
        f"/api/jobs/{job}/ai_output",
        json={
            "source_revision": extracted["source_revision"],
            "results": results,
            "summary": {
                "status": "partial",
                "total": 2,
                "succeeded": 1,
                "failed": 1,
                "cancelled": 0,
                "succeeded_ids": [0],
                "failed_ids": [1],
                "cancelled_ids": [],
            },
            "diagnostics": [{"finish_reason": "length"}],
        },
    )
    assert saved.status_code == 200, saved.text
    assert saved.json()["complete"] is False
    assert saved.json()["stage"] == "translation_partial"
    apply_response = _apply_current(
        client,
        job,
        extracted["source_revision"],
        saved.json()["translation_revision"],
    )
    assert apply_response.status_code == 409
    assert "不完整" in apply_response.json()["detail"]


def test_job_status_and_download(client, tmp_path):
    # 建任务 + 提取 + 翻译 + apply
    r = client.post("/api/jobs", json={"job": "dl", "input": str(tmp_path / "input" / "sample.xlsx")})
    assert r.status_code == 200
    extracted = _extract(client, "dl")
    saved = _save_translations(
        client,
        "dl",
        extracted["source_revision"],
        ["你好", "世界"],
    )
    applied = _apply_current(
        client,
        "dl",
        extracted["source_revision"],
        saved["translation_revision"],
        sep="\\n",
    )
    assert applied.status_code == 200, applied.text

    # 任务状态（列表带 stage/output）
    r = client.get("/api/jobs")
    dl = [j for j in r.json() if j["job"] == "dl"][0]
    assert dl["stage"] == "已导出"
    assert dl["output_translated"] and dl["output_translated"].endswith(".translated.xlsx")

    # 下载
    r = client.get("/api/jobs/dl/download", params={"kind": "translated"})
    assert r.status_code == 200
    assert r.content[:2] == b"PK"  # xlsx 是 zip
    r = client.get("/api/jobs/dl/download", params={"kind": "nope"})
    assert r.status_code == 404


@pytest.mark.p0_regression
def test_p0_05_reextract_invalidates_stale_outputs(client, app_workspace):
    """A new extraction revision must not inherit old translated/exported state."""
    job = "p0_reextract"
    created = client.post(
        "/api/jobs",
        json={"job": job, "input": str(app_workspace.sample_xlsx)},
    )
    assert created.status_code == 200, created.text
    extracted = _extract(client, job)
    saved = _save_translations(
        client,
        job,
        extracted["source_revision"],
        ["你好", "世界"],
    )
    applied = _apply_current(
        client,
        job,
        extracted["source_revision"],
        saved["translation_revision"],
    )
    assert applied.status_code == 200, applied.text

    copied_input = app_workspace.work_dir / job / "sample.xlsx"
    wb = openpyxl.load_workbook(copied_input)
    wb.active["A1"] = "Changed"
    wb.save(copied_input)

    reextracted = _extract(client, job)
    assert reextracted["source_revision"] != extracted["source_revision"]
    status = client.get(f"/api/jobs/{job}/status").json()
    assert status["stage"] == "已提取"
    assert status["output_translated"] is None
    assert status["output_bilingual"] is None
    assert client.get(
        f"/api/jobs/{job}/download",
        params={"kind": "translated"},
    ).status_code == 404
    stale_apply = _apply_current(
        client,
        job,
        extracted["source_revision"],
        saved["translation_revision"],
    )
    assert stale_apply.status_code == 409


@pytest.mark.p0_regression
def test_p0_05_cross_job_revision_cannot_be_saved(client, app_workspace):
    second_input = app_workspace.input_dir / "second.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "Different"
    wb.active["B1"] = "Content"
    wb.save(second_input)

    for job, input_path in (
        ("first", app_workspace.sample_xlsx),
        ("second", second_input),
    ):
        response = client.post(
            "/api/jobs",
            json={"job": job, "input": str(input_path)},
        )
        assert response.status_code == 200, response.text

    first = _extract(client, "first")
    second = _extract(client, "second")
    assert first["source_revision"] != second["source_revision"]

    response = client.post(
        "/api/jobs/second/ai_output",
        json={
            "source_revision": first["source_revision"],
            "results": [
                {"id": 0, "translation": "错误任务", "status": "succeeded"},
                {"id": 1, "translation": "错误内容", "status": "succeeded"},
            ],
            "summary": {
                "status": "succeeded",
                "total": 2,
                "succeeded": 2,
                "failed": 0,
                "cancelled": 0,
                "succeeded_ids": [0, 1],
                "failed_ids": [],
                "cancelled_ids": [],
            },
        },
    )
    assert response.status_code == 409
    assert client.get("/api/jobs/second/ai_output").json()["results"] == []


@pytest.mark.p0_regression
def test_p0_05_failed_reexport_keeps_previous_atomic_output(
    client,
    app_workspace,
    monkeypatch,
):
    job = "atomic_export"
    assert client.post(
        "/api/jobs",
        json={"job": job, "input": str(app_workspace.sample_xlsx)},
    ).status_code == 200
    extracted = _extract(client, job)
    saved = _save_translations(
        client,
        job,
        extracted["source_revision"],
        ["你好", "世界"],
    )
    first = _apply_current(
        client,
        job,
        extracted["source_revision"],
        saved["translation_revision"],
    )
    assert first.status_code == 200, first.text
    first_status = client.get(f"/api/jobs/{job}/status").json()

    import office_translate.jobs as jobs_module

    real_replace = jobs_module.os.replace

    def fail_bilingual_publish(source, target):
        if Path(target).name.endswith(".bilingual.xlsx"):
            raise OSError("simulated second output failure")
        return real_replace(source, target)

    monkeypatch.setattr(jobs_module.os, "replace", fail_bilingual_publish)
    failed = _apply_current(
        client,
        job,
        extracted["source_revision"],
        saved["translation_revision"],
        sep="\\r\\n",
    )
    assert failed.status_code == 400

    current = client.get(f"/api/jobs/{job}/status").json()
    assert current["stage_code"] == "exported"
    assert current["output_translated"] == first_status["output_translated"]
    assert current["output_bilingual"] == first_status["output_bilingual"]
    output_dir = app_workspace.work_dir / job / "output"
    assert sorted(path.name for path in output_dir.glob("*.xlsx")) == sorted(
        [
            Path(first_status["output_translated"]).name,
            Path(first_status["output_bilingual"]).name,
        ]
    )


def test_legacy_job_requires_reextract_and_discards_legacy_outputs(
    client,
    app_workspace,
):
    job = "legacy"
    assert client.post(
        "/api/jobs",
        json={"job": job, "input": str(app_workspace.sample_xlsx)},
    ).status_code == 200
    job_dir = app_workspace.work_dir / job
    (job_dir / "manifest.json").unlink()
    (job_dir / "source.txt").write_text("old", encoding="utf-8")
    (job_dir / "map.json").write_text("[]", encoding="utf-8")
    (job_dir / "translated.txt").write_text("old", encoding="utf-8")
    (job_dir / "ai_output.json").write_text("{}", encoding="utf-8")
    output_dir = job_dir / "output"
    output_dir.mkdir()
    legacy_output = output_dir / f"{job}_translated.xlsx"
    legacy_output.write_bytes(b"old")

    listed = next(item for item in client.get("/api/jobs").json() if item["job"] == job)
    assert listed["needs_reextract"] is True
    assert listed["stage"] == "需要重新提取"

    _extract(client, job)
    assert not (job_dir / "source.txt").exists()
    assert not (job_dir / "map.json").exists()
    assert not (job_dir / "translated.txt").exists()
    assert not (job_dir / "ai_output.json").exists()
    assert not legacy_output.exists()


def test_corrupt_manifest_can_be_recovered_by_explicit_reextract(
    client,
    app_workspace,
):
    job = "corrupt_manifest"
    assert client.post(
        "/api/jobs",
        json={"job": job, "input": str(app_workspace.sample_xlsx)},
    ).status_code == 200
    manifest = app_workspace.work_dir / job / "manifest.json"
    manifest.write_text("{broken", encoding="utf-8")

    listed = next(item for item in client.get("/api/jobs").json() if item["job"] == job)
    assert listed["stage_code"] == "error"
    assert listed["needs_reextract"] is True

    extracted = _extract(client, job)
    assert extracted["unique_texts"] == 2
    status = client.get(f"/api/jobs/{job}/status").json()
    assert status["stage_code"] == "extracted"
    assert status["source_revision"] == extracted["source_revision"]


def test_delete_job(client, tmp_path):
    client.post("/api/jobs", json={"job": "del1", "input": str(tmp_path / "input" / "sample.xlsx")})
    assert os.path.isdir(tmp_path / "work" / "del1")
    r = client.delete("/api/jobs/del1")
    assert r.json()["removed"] is True
    assert not os.path.isdir(tmp_path / "work" / "del1")


def test_glossary_flow(client):
    # 初始为空
    r = client.get("/api/glossary")
    assert r.json()["categories"] == {}

    # 新增
    r = client.post("/api/glossary/terms", json={"category": "汽车行业", "source": "PPB", "target": "十亿分之几"})
    assert r.status_code == 200, r.text
    r = client.get("/api/glossary")
    assert r.json()["categories"]["汽车行业"][0]["source"] == "PPB"

    # 删除
    r = client.delete("/api/glossary/terms", params={"category": "汽车行业", "source": "PPB"})
    assert r.json()["removed"] is True


def test_translate_google_mocked(client, monkeypatch):
    # mock GoogleProvider，避免真请求
    class FakeGoogle:
        name = "google"

        def __init__(self, mirrors):
            pass

        def translate(self, text, source, target):
            return f"译:{text}"

    import office_translate.gui.server as server_mod
    monkeypatch.setattr(server_mod, "GoogleProvider", FakeGoogle)

    r = client.post("/api/translate", json={"texts": ["Hello", "World"], "engine": "google"})
    assert r.status_code == 200, r.text
    results = r.json()["results"]
    assert results[0]["translation"] == "译:Hello"
    assert results[1]["translation"] == "译:World"


def test_translate_stream_sse(client, monkeypatch):
    """SSE ends with one validated succeeded summary."""
    import office_translate.gui.server as server_mod

    class FakeStreamGoogle:
        name = "google"

        def __init__(self, mirrors):
            pass

        def translate_stream(self, items, source, target):
            yield {
                "type": "block_succeeded",
                "items": [
                    {
                        "id": item.id,
                        "translation": f"译:{item.text}",
                        "uncertain_terms": [],
                    }
                    for item in reversed(items)
                ],
                "thinking": "",
                "diagnostic": {"finish_reason": "stop"},
            }

    monkeypatch.setattr(server_mod, "GoogleProvider", FakeStreamGoogle)

    r = client.post("/api/translate/stream", json={"texts": ["A", "B"], "engine": "google"})
    assert r.status_code == 200
    assert "text/event-stream" in r.headers.get("content-type", "")

    body = r.text
    assert '"type": "meta"' in body
    assert '"type": "progress"' in body
    assert '"type": "summary"' in body
    assert '"type": "done"' not in body
    assert '"type": "end"' not in body
    import json as _json

    translations = []
    summary = None
    for line in body.splitlines():
        if line.startswith("data: "):
            d = _json.loads(line[6:])
            if d.get("type") == "item_succeeded":
                translations.append(d.get("translation", ""))
            elif d.get("type") == "summary":
                summary = d
    assert "译:A" in translations and "译:B" in translations
    assert summary["status"] == "succeeded"
    assert summary["total"] == summary["succeeded"] == 2


def test_translate_stream_chunks(client, monkeypatch):
    """Every ID is emitted once and progress may reach 100 only with summary."""
    import office_translate.gui.server as server_mod

    class FakeStreamGoogle:
        name = "google"

        def __init__(self, mirrors):
            pass

        def translate_stream(self, items, source, target):
            yield {
                "type": "block_succeeded",
                "items": [
                    {"id": item.id, "translation": item.text, "uncertain_terms": []}
                    for item in items
                ],
                "thinking": "",
                "diagnostic": None,
            }

    monkeypatch.setattr(server_mod, "GoogleProvider", FakeStreamGoogle)

    texts = ["L1", "L2", "L3", "L4", "L5"]
    r = client.post("/api/translate/stream", json={"texts": texts, "engine": "google"})
    assert r.status_code == 200
    body = r.text
    assert body.count('"type": "item_succeeded"') == 5
    assert '"progress": 100' in body
    assert '"status": "succeeded"' in body


@pytest.mark.parametrize(
    ("output_format", "render"),
    [
        (
            "json",
            lambda items: '{"items":[' + ",".join(
                '{"id":%d,"translation":"译%d","uncertain_terms":[]}' % (item.id, item.id)
                for item in items
            ) + "]}",
        ),
        (
            "xml",
            lambda items: "<items>" + "".join(
                '<item id="%d"><translation>译%d</translation><uncertain_terms/></item>'
                % (item.id, item.id)
                for item in items
            ) + "</items>",
        ),
        (
            "text",
            lambda items: "".join("译%d\n" % item.id for item in items),
        ),
    ],
)
def test_stream_emits_item_preview_instead_of_raw_content(
    client, monkeypatch, output_format, render
):
    import json

    import office_translate.gui.server as server_mod

    class FakeOpenAI:
        def __init__(self, base_url, api_key, model, model_config=None):
            pass

        def translate_stream(self, items, source, target, glossary_entries=None):
            raw = render(list(items))
            step = 4
            for index in range(0, len(raw), step):
                yield {"type": "content", "delta": raw[index : index + step]}
            yield {
                "type": "block_succeeded",
                "items": [
                    {
                        "id": item.id,
                        "translation": f"译{item.id}",
                        "uncertain_terms": [],
                    }
                    for item in items
                ],
                "thinking": "",
                "diagnostic": {"finish_reason": "stop"},
            }

    monkeypatch.setattr(server_mod, "OpenAICompatProvider", FakeOpenAI)
    model_config = {"output_format": output_format}
    if output_format == "json":
        model_config["response_format"] = "none"
    response = client.post(
        "/api/translate/stream",
        json={
            "texts": ["A", "B"],
            "engine": "openai",
            "model_config": model_config,
        },
    )
    assert response.status_code == 200
    events = [
        json.loads(line[6:])
        for line in response.text.splitlines()
        if line.startswith("data: ")
    ]
    previews = [event for event in events if event.get("type") == "item_preview"]
    assert [event["id"] for event in previews] == [0, 1]
    assert [event["translation"] for event in previews] == ["译0", "译1"]
    assert not any(event.get("type") == "content" for event in events)
    succeeded = [event for event in events if event.get("type") == "item_succeeded"]
    assert [event["id"] for event in succeeded] == [0, 1]
    first_succeeded_index = next(
        index for index, event in enumerate(events)
        if event.get("type") == "item_succeeded"
    )
    assert all(
        next(index for index, candidate in enumerate(events) if candidate is event)
        < first_succeeded_index
        for event in previews
    )
    assert events[-1]["type"] == "summary"
    assert events[-1]["status"] == "succeeded"


@pytest.mark.parametrize(
    ("output_format", "content"),
    [
        (
            "json",
            '{"items":[{"id":0,"translation":"译0","uncertain_terms":[]},'
            '{"id":1,"translation":"unfinished',
        ),
        (
            "xml",
            "<items>"
            '<item id="0"><translation>译0</translation><uncertain_terms/></item>'
            '<item id="1"><translation>unfinished',
        ),
    ],
)
def test_stream_keeps_complete_preview_items_when_block_fails(
    client, monkeypatch, output_format, content
):
    """A failed block keeps structurally complete items; only missing ones fail."""
    import json

    import office_translate.gui.server as server_mod

    class PartialOpenAI:
        def __init__(self, base_url, api_key, model, model_config=None):
            pass

        def translate_stream(self, items, source, target, glossary_entries=None):
            step = 4
            for index in range(0, len(content), step):
                yield {"type": "content", "delta": content[index : index + step]}
            yield {
                "type": "block_failed",
                "ids": [item.id for item in items],
                "error_code": "provider_error",
                "error": "输出不完整",
                "thinking": "",
                "diagnostic": None,
            }

    monkeypatch.setattr(server_mod, "OpenAICompatProvider", PartialOpenAI)
    response = client.post(
        "/api/translate/stream",
        json={
            "texts": ["A", "B"],
            "engine": "openai",
            "model_config": {"output_format": output_format},
        },
    )
    events = [
        json.loads(line[6:])
        for line in response.text.splitlines()
        if line.startswith("data: ")
    ]
    succeeded = [event for event in events if event.get("type") == "item_succeeded"]
    failed = [event for event in events if event.get("type") == "item_failed"]
    assert [event["id"] for event in succeeded] == [0]
    assert [event["id"] for event in failed] == [1]
    summary = events[-1]
    assert summary["type"] == "summary"
    assert summary["status"] == "partial"
    assert summary["succeeded"] == 1
    assert summary["failed"] == 1


@pytest.mark.p0_regression
def test_p0_07_stream_rejects_missing_items(client, monkeypatch):
    """One returned item is kept; the missing item fails and summary is partial."""
    import json

    import office_translate.gui.server as server_mod

    class ShortStreamGoogle:
        name = "google"

        def __init__(self, mirrors):
            pass

        def translate_stream(self, items, source, target):
            yield {
                "type": "block_succeeded",
                "items": [
                    {"id": items[0].id, "translation": "ONLY", "uncertain_terms": []}
                ],
                "thinking": "",
                "diagnostic": {"raw_response": "short"},
            }

    monkeypatch.setattr(server_mod, "GoogleProvider", ShortStreamGoogle)
    response = client.post(
        "/api/translate/stream",
        json={"texts": ["A", "B"], "engine": "google"},
    )
    assert response.status_code == 200

    events = [
        json.loads(line[6:])
        for line in response.text.splitlines()
        if line.startswith("data: ")
    ]
    assert any(
        event.get("type") == "summary" and event.get("status") == "partial"
        for event in events
    )
    assert [event["id"] for event in events if event.get("type") == "item_succeeded"] == [0]
    assert [event["id"] for event in events if event.get("type") == "item_failed"] == [1]


@pytest.mark.p0_regression
def test_stream_preserves_internal_newline_as_one_item(client, monkeypatch):
    import json
    import office_translate.gui.server as server_mod

    class NewlineProvider:
        name = "google"

        def __init__(self, mirrors):
            pass

        def translate_stream(self, items, source, target):
            yield {
                "type": "block_succeeded",
                "items": [
                    {"id": items[0].id, "translation": "甲\n乙", "uncertain_terms": []},
                    {"id": items[1].id, "translation": "丙", "uncertain_terms": []},
                ],
                "thinking": "",
                "diagnostic": None,
            }

    monkeypatch.setattr(server_mod, "GoogleProvider", NewlineProvider)
    response = client.post(
        "/api/translate/stream",
        json={"texts": ["A\nB", "C"], "engine": "google"},
    )
    events = [
        json.loads(line[6:])
        for line in response.text.splitlines()
        if line.startswith("data: ")
    ]
    succeeded = [event for event in events if event.get("type") == "item_succeeded"]
    assert [(item["id"], item["translation"]) for item in succeeded] == [
        (0, "甲\n乙"),
        (1, "丙"),
    ]
    assert events[-1]["type"] == "summary"
    assert events[-1]["status"] == "succeeded"


@pytest.mark.p0_regression
def test_stream_rejects_legacy_done_event(client, monkeypatch):
    import json
    import office_translate.gui.server as server_mod

    class LegacyProvider:
        name = "google"

        def __init__(self, mirrors):
            pass

        def translate_stream(self, items, source, target):
            yield {"type": "done", "id": 0, "translation": "legacy"}

    monkeypatch.setattr(server_mod, "GoogleProvider", LegacyProvider)
    response = client.post(
        "/api/translate/stream",
        json={"texts": ["A"], "engine": "google"},
    )
    events = [
        json.loads(line[6:])
        for line in response.text.splitlines()
        if line.startswith("data: ")
    ]
    assert events[-1]["type"] == "summary"
    assert events[-1]["status"] == "failed"
    assert not any(event.get("type") == "item_succeeded" for event in events)


@pytest.mark.p0_regression
def test_stream_exception_still_emits_failed_item_and_summary(client, monkeypatch):
    import json
    import office_translate.gui.server as server_mod

    class InterruptedProvider:
        name = "google"

        def __init__(self, mirrors):
            pass

        def translate_stream(self, items, source, target):
            yield {"type": "content", "delta": '{"items":['}
            raise ConnectionError("socket closed")

    monkeypatch.setattr(server_mod, "GoogleProvider", InterruptedProvider)
    response = client.post(
        "/api/translate/stream",
        json={"texts": ["A"], "engine": "google"},
    )
    events = [
        json.loads(line[6:])
        for line in response.text.splitlines()
        if line.startswith("data: ")
    ]
    failed = next(event for event in events if event.get("type") == "item_failed")
    assert failed["error_code"] == "provider_error"
    assert "socket closed" not in failed["error"]
    assert events[-1]["type"] == "summary"
    assert events[-1]["status"] == "failed"


def test_stream_rejects_malformed_failure_event(client, monkeypatch):
    import json
    import office_translate.gui.server as server_mod

    class MalformedProvider:
        name = "google"

        def __init__(self, mirrors):
            pass

        def translate_stream(self, items, source, target):
            yield {
                "type": "block_failed",
                "ids": [items[0].id],
                "error_code": "provider_error",
                "error": {"unexpected": "object"},
                "thinking": "",
            }

    monkeypatch.setattr(server_mod, "GoogleProvider", MalformedProvider)
    response = client.post(
        "/api/translate/stream",
        json={"texts": ["A"], "engine": "google"},
    )
    events = [
        json.loads(line[6:])
        for line in response.text.splitlines()
        if line.startswith("data: ")
    ]
    failed = next(event for event in events if event.get("type") == "item_failed")
    assert failed["error_code"] == "invalid_provider_event"
    assert events[-1]["status"] == "failed"


def test_stream_reassembles_long_source_without_exposing_segment_ids(client, monkeypatch):
    """Long source cells are translated in segments but remain one GUI row."""
    import json
    import office_translate.gui.server as server_mod

    class SegmentedOpenAI:
        def __init__(self, base_url, api_key, model, model_config=None):
            pass

        def translate_stream(self, items, source, target, glossary_entries=None):
            for item in items:
                yield {
                    "type": "content",
                    "delta": (
                        '<items><item id="%d"><translation>%s</translation>'
                        "<uncertain_terms/></item></items>"
                    )
                    % (item.id, "译" + str(item.id)),
                }
                yield {
                    "type": "block_succeeded",
                    "items": [
                        {
                            "id": item.id,
                            "translation": "译" + str(item.id),
                            "uncertain_terms": [],
                        }
                    ],
                    "thinking": "",
                    "diagnostic": None,
                }

    monkeypatch.setattr(server_mod, "OpenAICompatProvider", SegmentedOpenAI)
    response = client.post(
        "/api/translate/stream",
        json={
            "texts": ["x" * 1600],
            "engine": "openai",
            "model_config": {
                "output_format": "xml",
                "model_context": 1000,
                "max_tokens": 100,
            },
        },
    )
    assert response.status_code == 200, response.text
    events = [
        json.loads(line[6:])
        for line in response.text.splitlines()
        if line.startswith("data: ")
    ]
    previews = [event for event in events if event["type"] == "item_preview"]
    terminal = [
        event
        for event in events
        if event["type"] in {"item_succeeded", "item_failed", "item_cancelled"}
    ]
    assert previews and all(event["id"] == 0 for event in previews)
    assert [event["id"] for event in terminal] == [0]
    assert terminal[0]["type"] == "item_succeeded"
    assert terminal[0]["translation"].startswith("译")
    assert events[-1]["status"] == "succeeded"


def test_stream_concurrency_is_bounded(client, monkeypatch):
    import json
    import threading
    import time
    import office_translate.gui.server as server_mod

    active = 0
    peak = 0
    lock = threading.Lock()

    class ConcurrentGoogle:
        name = "google"

        def __init__(self, mirrors):
            pass

        def translate_stream(self, items, source, target):
            nonlocal active, peak
            with lock:
                active += 1
                peak = max(peak, active)
            time.sleep(0.04)
            yield {
                "type": "block_succeeded",
                "items": [
                    {"id": item.id, "translation": "译", "uncertain_terms": []}
                    for item in items
                ],
                "thinking": "",
                "diagnostic": None,
            }
            with lock:
                active -= 1

    monkeypatch.setattr(server_mod, "GoogleProvider", ConcurrentGoogle)
    response = client.post(
        "/api/translate/stream",
        json={
            "texts": ["x" * 4000 for _ in range(6)],
            "engine": "google",
            "concurrency": 2,
        },
    )
    assert response.status_code == 200, response.text
    events = [
        json.loads(line[6:])
        for line in response.text.splitlines()
        if line.startswith("data: ")
    ]
    assert peak == 2
    assert events[-1]["status"] == "succeeded"
    assert events[-1]["total"] == 6


def test_review_is_stable_aggregated_and_gates_export(client, app_workspace):
    job = "review_persistence"
    assert client.post(
        "/api/jobs",
        json={"job": job, "input": str(app_workspace.sample_xlsx)},
    ).status_code == 200
    extracted = _extract(client, job)
    results = [
        {
            "id": 0,
            "translation": "API 调用",
            "status": "succeeded",
            "uncertain_terms": [
                {"term": "API", "reason": "术语", "candidate": "应用程序接口"}
            ],
        },
        {
            "id": 1,
            "translation": "API 文档",
            "status": "succeeded",
            "uncertain_terms": [
                {"term": "API", "reason": "术语", "candidate": "应用程序接口"}
            ],
        },
    ]
    saved = client.post(
        f"/api/jobs/{job}/ai_output",
        json={
            "source_revision": extracted["source_revision"],
            "results": results,
            "summary": {
                "status": "succeeded",
                "total": 2,
                "succeeded": 2,
                "failed": 0,
                "cancelled": 0,
                "succeeded_ids": [0, 1],
                "failed_ids": [],
                "cancelled_ids": [],
            },
        },
    ).json()
    first = client.get(
        f"/api/jobs/{job}/review",
        params={
            "source_revision": extracted["source_revision"],
            "translation_revision": saved["translation_revision"],
        },
    ).json()
    second = client.get(
        f"/api/jobs/{job}/review",
        params={"source_revision": extracted["source_revision"]},
    ).json()
    assert len(first["items"]) == 1
    assert first["items"][0]["row_ids"] == [0, 1]
    assert first["items"][0]["selected_row_ids"] == [0, 1]
    assert first["items"][0]["review_id"] == second["items"][0]["review_id"]
    blocked = _apply_current(
        client,
        job,
        extracted["source_revision"],
        saved["translation_revision"],
    )
    assert blocked.status_code == 409
    decision = {
        "source_revision": extracted["source_revision"],
        "translation_revision": saved["translation_revision"],
        "decisions": [
            {
                "review_id": first["items"][0]["review_id"],
                "kind": "term",
                "decision": "edited",
                "target": "接口",
                "category": "软件",
                "apply_to_text": True,
                "empty_translation_confirmed": False,
                "row_ids": [0, 1],
                "selected_row_ids": [1],
            }
        ],
    }
    invalid_scope = {
        **decision,
        "decisions": [
            {
                **decision["decisions"][0],
                "selected_row_ids": [99],
            }
        ],
    }
    rejected = client.put(f"/api/jobs/{job}/review", json=invalid_scope)
    assert rejected.status_code == 409
    assert "超出" in rejected.json()["detail"]
    duplicate_scope = {
        **decision,
        "decisions": [
            {
                **decision["decisions"][0],
                "selected_row_ids": [1, 1],
            }
        ],
    }
    rejected_duplicate = client.put(f"/api/jobs/{job}/review", json=duplicate_scope)
    assert rejected_duplicate.status_code == 409
    assert "不能重复" in rejected_duplicate.json()["detail"]
    missing_selection = {
        **decision,
        "decisions": [
            {
                key: value
                for key, value in decision["decisions"][0].items()
                if key != "selected_row_ids"
            }
        ],
    }
    rejected_missing = client.put(f"/api/jobs/{job}/review", json=missing_selection)
    assert rejected_missing.status_code == 409
    assert "selected_row_ids 必须" in rejected_missing.json()["detail"]
    updated = client.put(f"/api/jobs/{job}/review", json=decision)
    assert updated.status_code == 200, updated.text
    review = updated.json()
    assert review["items"][0]["decision"] == "edited"
    assert review["items"][0]["target"] == "接口"
    assert review["items"][0]["selected_row_ids"] == [1]
    output = client.get(f"/api/jobs/{job}/ai_output").json()
    assert [item["translation"] for item in output["results"]] == [
        "API 调用",
        "接口 文档",
    ]
    applied = _apply_current(
        client,
        job,
        extracted["source_revision"],
        output["translation_revision"],
    )
    assert applied.status_code == 200, applied.text
    stale_source = client.get(
        f"/api/jobs/{job}/review",
        params={"source_revision": "0" * 64},
    )
    assert stale_source.status_code == 409
    stale_translation = client.get(
        f"/api/jobs/{job}/review",
        params={
            "source_revision": extracted["source_revision"],
            "translation_revision": "0" * 64,
        },
    )
    assert stale_translation.status_code == 409


def test_blank_translation_requires_explicit_review_confirmation(client, app_workspace):
    job = "blank_review"
    assert client.post(
        "/api/jobs",
        json={"job": job, "input": str(app_workspace.sample_xlsx)},
    ).status_code == 200
    extracted = _extract(client, job)
    saved_response = client.post(
        f"/api/jobs/{job}/ai_output",
        json={
            "source_revision": extracted["source_revision"],
            "results": [
                {"id": 0, "translation": "", "status": "succeeded"},
                {"id": 1, "translation": "世界", "status": "succeeded"},
            ],
            "summary": {
                "status": "succeeded",
                "total": 2,
                "succeeded": 2,
                "failed": 0,
                "cancelled": 0,
                "succeeded_ids": [0, 1],
                "failed_ids": [],
                "cancelled_ids": [],
            },
        },
    )
    saved = saved_response.json()
    review = client.get(
        f"/api/jobs/{job}/review",
        params={"source_revision": extracted["source_revision"]},
    ).json()
    blank = next(item for item in review["items"] if item["kind"] == "blank_translation")
    blocked = _apply_current(
        client,
        job,
        extracted["source_revision"],
        saved["translation_revision"],
    )
    assert blocked.status_code == 409
    accepted = client.put(
        f"/api/jobs/{job}/review",
        json={
            "source_revision": extracted["source_revision"],
            "translation_revision": saved["translation_revision"],
            "decisions": [
                {
                    "review_id": blank["review_id"],
                    "kind": "blank_translation",
                    "decision": "accepted",
                    "target": "",
                    "category": None,
                    "apply_to_text": False,
                    "empty_translation_confirmed": True,
                    "row_ids": [0],
                    "selected_row_ids": [],
                }
            ],
        },
    )
    assert accepted.status_code == 200, accepted.text
    output = client.get(f"/api/jobs/{job}/ai_output").json()
    assert output["complete"] is True
    applied = _apply_current(
        client,
        job,
        extracted["source_revision"],
        output["translation_revision"],
    )
    assert applied.status_code == 200, applied.text


def test_translate_unknown_engine(client):
    r = client.post("/api/translate", json={"texts": ["Hello"], "engine": "nope"})
    assert r.status_code == 400


def test_translate_rejects_invalid_response_format_combination(client):
    r = client.post(
        "/api/translate",
        json={
            "texts": ["Hello"],
            "engine": "openai",
            "model_config": {"output_format": "text", "response_format": "json_schema"},
        },
    )
    assert r.status_code == 400


def test_translate_stream_rejects_invalid_response_format_combinations(client):
    r = client.post(
        "/api/translate/stream",
        json={
            "texts": ["A"],
            "engine": "openai",
            "model_config": {"output_format": "xml", "response_format": "json_object"},
        },
    )
    assert r.status_code == 400
    r = client.post(
        "/api/translate/stream",
        json={
            "texts": ["A"],
            "engine": "openai",
            "model_config": {"output_format": "json", "response_format": "yaml"},
        },
    )
    assert r.status_code == 400


def test_glossary_edit_and_delete(client):
    # 新增
    r = client.post("/api/glossary/terms", json={"category": "软件", "source": "API", "target": "接口"})
    assert r.status_code == 200

    # 编辑
    r = client.put("/api/glossary/terms", json={"category": "软件", "source": "API", "target": "应用程序接口", "note": "更新"})
    assert r.status_code == 200
    assert r.json()["target"] == "应用程序接口"

    # 编辑不存在的 → 404
    r = client.put("/api/glossary/terms", json={"category": "软件", "source": "不存在", "target": "x"})
    assert r.status_code == 404

    # 删除
    r = client.delete("/api/glossary/terms", params={"category": "软件", "source": "API"})
    assert r.json()["removed"] is True


def test_mirrors_test_mocked(client, monkeypatch):
    import office_translate.gui.server as server_mod

    class FakeGoogle:
        name = "google"

        def __init__(self, mirrors):
            pass

        def translate(self, text, source, target):
            return "ok"

    monkeypatch.setattr(server_mod, "GoogleProvider", FakeGoogle)
    r = client.post("/api/mirrors/test", json={"mirrors": ["https://a.example", "https://b.example"]})
    assert r.status_code == 200, r.text
    results = r.json()["results"]
    assert len(results) == 2
    assert all(x["ok"] for x in results)
    assert results[0]["url"] == "https://a.example"


def test_mirrors_default(client):
    r = client.get("/api/mirrors")
    assert r.status_code == 200
    assert len(r.json()["mirrors"]) == 3


def test_settings_roundtrip(client):
    # 读取默认
    r = client.get("/api/settings")
    assert r.status_code == 200
    settings = r.json()
    assert settings["ai"]["engine"] == "google"
    assert "openai" in settings["ai"]["providers"]

    # 修改镜像站 + 引擎
    settings["ai"]["mirrors"] = ["https://custom.example"]
    settings["ai"]["engine"] = "openai"
    r = client.put("/api/settings", json=settings)
    assert r.status_code == 200
    assert r.json()["ai"]["mirrors"] == ["https://custom.example"]

    # 重新读取（持久化）
    r = client.get("/api/settings")
    assert r.json()["ai"]["mirrors"] == ["https://custom.example"]


def test_settings_reject_missing_active_provider_or_model(client):
    settings = client.get("/api/settings").json()
    original_provider = settings["ai"]["active_provider"]
    del settings["ai"]["providers"][original_provider]
    missing_provider = client.put("/api/settings", json=settings)
    assert missing_provider.status_code == 400

    settings = client.get("/api/settings").json()
    settings["ai"]["active_model"] = "deleted-model"
    missing_model = client.put("/api/settings", json=settings)
    assert missing_model.status_code == 400
    assert client.get("/api/settings").json()["ai"]["active_provider"] == original_provider

def test_settings_api_never_returns_provider_secret(client, app_workspace):
    settings = client.get("/api/settings").json()
    provider = settings["ai"]["providers"]["openai"]
    assert "api_key" not in provider
    assert provider["api_key_configured"] is False

    provider["api_key"] = "secret-key-for-test"
    saved = client.put("/api/settings", json=settings)
    assert saved.status_code == 200, saved.text
    saved_provider = saved.json()["ai"]["providers"]["openai"]
    assert "api_key" not in saved_provider
    assert saved_provider["api_key_configured"] is True
    assert "secret-key-for-test" not in saved.text

    loaded = client.get("/api/settings")
    assert loaded.status_code == 200
    loaded_provider = loaded.json()["ai"]["providers"]["openai"]
    assert "api_key" not in loaded_provider
    assert loaded_provider["api_key_masked"].endswith("test")
    assert "secret-key-for-test" not in loaded.text

    raw_settings = (app_workspace.root / "data" / "gui_settings.json").read_text(
        encoding="utf-8"
    )
    assert "secret-key-for-test" in raw_settings


def test_translation_resolves_saved_provider_secret_server_side(client, monkeypatch):
    import office_translate.gui.server as server_mod

    settings = client.get("/api/settings").json()
    settings["ai"]["providers"]["openai"]["api_key"] = "server-only-secret"
    assert client.put("/api/settings", json=settings).status_code == 200

    captured = {}

    class CapturingProvider:
        def __init__(self, base_url, api_key, model, model_config=None):
            captured.update(
                base_url=base_url,
                api_key=api_key,
                model=model,
                model_config=model_config,
            )

    def fake_translate_batch(texts, provider, source, target, matched):
        return [
            {
                "id": item_id,
                "translation": f"译{text}",
                "uncertain_terms": [],
                "status": "succeeded",
                "error": None,
            }
            for item_id, text in enumerate(texts)
        ]

    monkeypatch.setattr(server_mod, "OpenAICompatProvider", CapturingProvider)
    monkeypatch.setattr(server_mod, "translate_batch", fake_translate_batch)
    response = client.post(
        "/api/translate",
        json={
            "texts": ["Hello"],
            "engine": "openai",
            "provider_config": {
                "provider_id": "openai",
                "model": "gpt-4o-mini",
                "api_key": "browser-override-must-be-ignored",
            },
            "model_config": {"output_format": "xml"},
        },
    )
    assert response.status_code == 200, response.text
    assert captured["api_key"] == "server-only-secret"
    assert captured["api_key"] != "browser-override-must-be-ignored"
    assert "server-only-secret" not in response.text


def test_gui_does_not_grant_wildcard_cors(client):
    response = client.get("/api/jobs", headers={"Origin": "http://untrusted.example"})
    assert response.status_code == 200
    assert "access-control-allow-origin" not in response.headers


def test_gui_root_serves_local_runtime_and_diagnostic_log(app, client):
    page = client.get("/")
    assert page.status_code == 200
    assert 'src="vendor/vue.global.prod.js"' in page.text
    assert "unpkg.com" not in page.text

    log_path = Path(app.state.diagnostic_log_path)
    assert log_path.is_file()
    assert "api_key" not in log_path.read_text(encoding="utf-8").lower()


def test_glossary_category_delete(client):
    client.post("/api/glossary/terms", json={"category": "甲", "source": "A", "target": "a"})
    client.post("/api/glossary/terms", json={"category": "甲", "source": "B", "target": "b"})
    client.post("/api/glossary/terms", json={"category": "乙", "source": "C", "target": "c"})

    r = client.delete("/api/glossary/categories", params={"category": "甲"})
    assert r.json() == {"removed": True, "count": 2}
    g = client.get("/api/glossary").json()
    assert "甲" not in g["categories"]
    assert "乙" in g["categories"]


def test_glossary_batch_delete(client):
    client.post("/api/glossary/terms", json={"category": "丙", "source": "X", "target": "x"})
    client.post("/api/glossary/terms", json={"category": "丙", "source": "Y", "target": "y"})
    r = client.post("/api/glossary/batch-delete", json={"category": "丙", "sources": ["X"]})
    assert r.json() == {"removed": 1}
    g = client.get("/api/glossary").json()
    assert [e["source"] for e in g["categories"]["丙"]] == ["Y"]
