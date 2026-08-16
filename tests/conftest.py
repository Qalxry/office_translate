"""Shared isolated workspaces and workbook fixtures for the test suite."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from office_translate.gui.server import create_app


@dataclass(frozen=True)
class GuiWorkspace:
    """Paths owned by one GUI test invocation."""

    root: Path
    config_path: Path
    input_dir: Path
    work_dir: Path
    glossary_path: Path
    sample_xlsx: Path


@pytest.fixture
def app_workspace(tmp_path: Path) -> GuiWorkspace:
    """Create an isolated config, input workbook, work dir, and glossary."""
    config_path = tmp_path / "config.yaml"
    config_path.write_text(
        "work_dir: work\noutput_dir: output\nsep: '\\n'\n",
        encoding="utf-8",
    )

    input_dir = tmp_path / "input"
    input_dir.mkdir()
    sample_xlsx = input_dir / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Hello"
    ws["B1"] = "World"
    wb.save(sample_xlsx)

    return GuiWorkspace(
        root=tmp_path,
        config_path=config_path,
        input_dir=input_dir,
        work_dir=tmp_path / "work",
        glossary_path=tmp_path / "glossary.json",
        sample_xlsx=sample_xlsx,
    )


@pytest.fixture
def app(app_workspace: GuiWorkspace):
    """Build a FastAPI app bound only to the current test workspace."""
    return create_app(
        config_path=str(app_workspace.config_path),
        glossary_path=str(app_workspace.glossary_path),
    )


@pytest.fixture
def client(app):
    """Yield a TestClient whose lifespan cannot leak into another test."""
    with TestClient(app) as test_client:
        yield test_client


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    """Create the shared round-trip workbook used by XLSX adapter tests."""
    path = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Hello"
    ws["B1"] = "World"
    ws["A2"] = "重复文本"
    ws["B2"] = "重复文本"
    ws["A3"] = 42
    ws["A4"] = "=SUM(1,2)"
    ws["A5"] = "Merged"
    ws.merge_cells("A5:B5")
    ws["A1"].font = openpyxl.styles.Font(bold=True, color="FF0000")
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Hello"
    wb.save(path)
    return path
