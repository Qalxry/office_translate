"""XLSX structured-artifact round-trip regression tests."""

import openpyxl
import pytest

from office_translate.artifacts import TranslationArtifact
from office_translate.formats.xlsx.applier import (
    EXCEL_CELL_TEXT_LIMIT,
    TranslationError,
    apply_artifacts,
)
from office_translate.formats.xlsx.extractor import extract_artifact


TRANSLATIONS = {
    "Hello": "你好",
    "World": "世界",
    "重复文本": "重复译文",
    "Merged": "合并",
}


def _translated_artifact(source):
    return TranslationArtifact.create(
        source,
        [
            {
                "id": item.id,
                "translation": TRANSLATIONS[item.text],
                "status": "succeeded",
            }
            for item in source.items
        ],
    )


def test_extract_artifact(sample_xlsx):
    source = extract_artifact(sample_xlsx)

    assert source.stats == {
        "sheets": ["Sheet1", "Sheet2"],
        "cells_total": 8,
        "cells_translatable": 6,
        "unique_texts": 4,
    }
    assert [item.text for item in source.items] == [
        "Hello",
        "World",
        "重复文本",
        "Merged",
    ]
    assert [cell.to_dict() for cell in source.items[0].cells] == [
        {"sheet": "Sheet1", "coordinate": "A1"},
        {"sheet": "Sheet2", "coordinate": "A1"},
    ]


def test_structured_roundtrip(sample_xlsx, tmp_path):
    source = extract_artifact(sample_xlsx)
    translation = _translated_artifact(source)
    out_t = tmp_path / "out_translated.xlsx"
    out_b = tmp_path / "out_bilingual.xlsx"

    result = apply_artifacts(
        sample_xlsx,
        source,
        translation,
        out_t,
        out_b,
        sep="\n",
    )
    assert result == {"unique_texts": 4, "cells_filled": 6}

    wb = openpyxl.load_workbook(out_t)
    ws = wb["Sheet1"]
    assert ws["A1"].value == "你好"
    assert ws["B1"].value == "世界"
    assert ws["A2"].value == "重复译文"
    assert ws["B2"].value == "重复译文"
    assert ws["A3"].value == 42
    assert ws["A4"].value == "=SUM(1,2)"
    assert ws["A5"].value == "合并"
    assert wb["Sheet2"]["A1"].value == "你好"
    assert ws["A1"].font.bold is True
    assert ws["A1"].font.color.rgb in ("FF0000", "00FF0000")
    assert any(str(cell_range) == "A5:B5" for cell_range in ws.merged_cells.ranges)

    wb2 = openpyxl.load_workbook(out_b)
    ws2 = wb2["Sheet1"]
    assert ws2["A1"].value == "Hello\n你好"
    assert ws2["B1"].value == "World\n世界"


@pytest.mark.p0_regression
def test_p0_05_apply_rejects_changed_source_cell(sample_xlsx, tmp_path):
    source = extract_artifact(sample_xlsx)
    translation = _translated_artifact(source)

    wb = openpyxl.load_workbook(sample_xlsx)
    wb["Sheet1"]["A1"] = "Changed after extraction"
    wb.save(sample_xlsx)

    with pytest.raises(TranslationError, match="摘要|原文|修订|revision"):
        apply_artifacts(
            sample_xlsx,
            source,
            translation,
            tmp_path / "changed-source.xlsx",
            tmp_path / "changed-source-bilingual.xlsx",
        )


@pytest.mark.p0_regression
def test_p0_06_structured_translations_preserve_control_characters(tmp_path):
    source_path = tmp_path / "control-characters.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for row, value in enumerate(
        ["path", "literal escape", "LF", "CR", "CRLF", "trailing slash"],
        start=1,
    ):
        ws.cell(row=row, column=1, value=value)
    wb.save(source_path)

    expected = [
        r"C:\new\report",
        r"literal\n",
        "line 1\nline 2",
        "carriage\rreturn",
        "windows\r\nline",
        "trailing\\",
    ]
    source = extract_artifact(source_path)
    translation = TranslationArtifact.create(
        source,
        [
            {"id": item_id, "translation": value}
            for item_id, value in enumerate(expected)
        ],
    )
    output = tmp_path / "structured.xlsx"
    bilingual = tmp_path / "structured-bilingual.xlsx"
    apply_artifacts(
        source_path,
        source,
        translation,
        output,
        bilingual,
        sep="\r\n",
    )

    translated_wb = openpyxl.load_workbook(output)
    assert [
        translated_wb.active.cell(row=row, column=1).value
        for row in range(1, 7)
    ] == expected
    bilingual_wb = openpyxl.load_workbook(bilingual)
    assert bilingual_wb.active["A1"].value == f"path\r\n{expected[0]}"


def test_incomplete_translation_artifact_cannot_be_exported(sample_xlsx, tmp_path):
    source = extract_artifact(sample_xlsx)
    translation = TranslationArtifact.create(
        source,
        [
            {"id": 0, "translation": "你好", "status": "succeeded"},
            {"id": 1, "translation": "World", "status": "failed"},
            {"id": 2, "translation": "重复译文", "status": "succeeded"},
            {"id": 3, "translation": "合并", "status": "succeeded"},
        ],
    )

    with pytest.raises(TranslationError, match="不完整|失败"):
        apply_artifacts(
            sample_xlsx,
            source,
            translation,
            tmp_path / "incomplete.xlsx",
            tmp_path / "incomplete-bilingual.xlsx",
        )


@pytest.mark.parametrize(
    ("translation_length", "should_export"),
    [(32_766, True), (32_767, False), (32_768, False)],
)
def test_excel_cell_text_limit_is_checked_before_publication(
    tmp_path,
    translation_length,
    should_export,
):
    source_path = tmp_path / "limit.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "s"
    wb.save(source_path)
    source = extract_artifact(source_path)
    translation = TranslationArtifact.create(
        source,
        [{"id": 0, "translation": "x" * translation_length}],
    )
    translated = tmp_path / "limit-translated.xlsx"
    bilingual = tmp_path / "limit-bilingual.xlsx"

    if should_export:
        apply_artifacts(source_path, source, translation, translated, bilingual, sep="")
        assert len(openpyxl.load_workbook(translated).active["A1"].value) == translation_length
    else:
        with pytest.raises(TranslationError) as exc_info:
            apply_artifacts(source_path, source, translation, translated, bilingual)
        assert any(
            diagnostic.code == "xlsx.cell_text_limit"
            and diagnostic.details["max_length"] == EXCEL_CELL_TEXT_LIMIT
            for diagnostic in exc_info.value.diagnostics
        )
        assert not translated.exists()
        assert not bilingual.exists()


def test_bilingual_separator_overhead_is_checked(tmp_path):
    source_path = tmp_path / "bilingual-limit.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "o" * (EXCEL_CELL_TEXT_LIMIT - 1)
    wb.save(source_path)
    source = extract_artifact(source_path)
    translation = TranslationArtifact.create(source, [{"id": 0, "translation": "t"}])
    translated = tmp_path / "bilingual-translated.xlsx"
    bilingual = tmp_path / "bilingual-bilingual.xlsx"

    with pytest.raises(TranslationError, match="预检") as exc_info:
        apply_artifacts(
            source_path,
            source,
            translation,
            translated,
            bilingual,
            sep="\n",
        )
    assert any(
        diagnostic.details.get("mode") == "bilingual"
        for diagnostic in exc_info.value.diagnostics
    )
    assert not translated.exists()
    assert not bilingual.exists()
