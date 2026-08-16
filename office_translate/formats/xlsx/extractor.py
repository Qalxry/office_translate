"""Extract validated, revision-bound source items from an XLSX workbook.

The GUI only accepts OOXML ``.xlsx`` files.  ``preflight_workbook`` is kept
separate from extraction so the upload/task layer can validate a file before
publishing a task directory.  A preflight warning is deliberately not hidden:
external links, protection and OOXML objects which are not covered by the
round-trip contract are returned as structured diagnostics for the GUI.
"""

from __future__ import annotations

import os
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.rich_text import CellRichText

from ...artifacts import ArtifactError, CellRef, SourceArtifact, SourceItem, sha256_file
from ...escape import escape_text
from ...storage import atomic_write_json, atomic_write_text


@dataclass(frozen=True)
class WorkbookDiagnostic:
    """A machine-readable workbook warning or blocking validation error."""

    code: str
    severity: str
    message: str
    action: str
    sheet: str | None = None
    coordinate: str | None = None
    details: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        result: dict[str, Any] = {
            "code": self.code,
            "severity": self.severity,
            "message": self.message,
            "action": self.action,
        }
        if self.sheet is not None:
            result["sheet"] = self.sheet
        if self.coordinate is not None:
            result["coordinate"] = self.coordinate
        if self.details:
            result["details"] = self.details
        return result


@dataclass(frozen=True)
class WorkbookPreflight:
    """The result of validating one candidate input before task creation."""

    path: str
    kind: str
    ok: bool
    diagnostics: tuple[WorkbookDiagnostic, ...] = ()

    @property
    def blocking_diagnostics(self) -> tuple[WorkbookDiagnostic, ...]:
        return tuple(
            diagnostic
            for diagnostic in self.diagnostics
            if diagnostic.severity == "error"
        )

    def to_dict(self) -> dict[str, Any]:
        return {
            "path": self.path,
            "kind": self.kind,
            "ok": self.ok,
            "diagnostics": [diagnostic.to_dict() for diagnostic in self.diagnostics],
        }


class WorkbookPreflightError(ArtifactError):
    """Raised by extraction when a workbook cannot be safely opened."""

    def __init__(self, result: WorkbookPreflight) -> None:
        self.result = result
        first = result.blocking_diagnostics[0] if result.blocking_diagnostics else None
        message = first.message if first is not None else "工作簿未通过预检"
        super().__init__(message)


@dataclass
class _TextGroup:
    text: str
    cells: list[CellRef] = field(default_factory=list)
    id: int = -1


_UNVERIFIED_OBJECT_PREFIXES = (
    "xl/activeX/",
    "xl/charts/",
    "xl/controlProps/",
    "xl/drawings/",
    "xl/embeddings/",
    "xl/externalLinks/",
    "xl/ink/",
    "xl/media/",
    "xl/model/",
    "xl/people/",
    "xl/pivotCache/",
    "xl/pivotTables/",
    "xl/queryTables/",
    "xl/richData/",
    "xl/slicer/",
    "xl/tables/",
    "xl/threadedComments/",
    "customXml/",
)


def cell_text(value: Any) -> str:
    """Return the visible text of a normal or rich-text cell value."""
    if isinstance(value, CellRichText):
        return "".join(
            part if isinstance(part, str) else getattr(part, "text", str(part))
            for part in value
        )
    return value if isinstance(value, str) else str(value)


def is_multi_run_rich_text(value: Any) -> bool:
    """Whether a value contains multiple inline rich-text runs."""
    return isinstance(value, CellRichText) and len(value) > 1


def _is_translatable(cell: Any) -> bool:
    value = cell.value
    if isinstance(value, CellRichText):
        return bool(cell_text(value))
    if cell.data_type != "s":
        return False
    return isinstance(value, str) and value != "" and not value.startswith("=")


def _rich_text_diagnostics(wb: Any) -> list[WorkbookDiagnostic]:
    diagnostics: list[WorkbookDiagnostic] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not is_multi_run_rich_text(cell.value):
                    continue
                value = cell.value
                diagnostics.append(
                    WorkbookDiagnostic(
                        code="xlsx.rich_text_runs",
                        severity="warning",
                        message=(
                            f"{ws.title}!{cell.coordinate} 含 {len(value)} 个富文本 run，"
                            "默认导出会阻止该单元格被扁平化。"
                        ),
                        action=(
                            "导出时明确选择 preserve_original 保留原文，或 flatten"
                            "（确认接受局部格式丢失）。"
                        ),
                        sheet=ws.title,
                        coordinate=cell.coordinate,
                        details={
                            "run_count": len(value),
                            "text_length": len(cell_text(value)),
                        },
                    )
                )
    return diagnostics


def _object_part_diagnostics(names: list[str]) -> list[WorkbookDiagnostic]:
    parts = sorted(
        name
        for name in names
        if name.startswith(_UNVERIFIED_OBJECT_PREFIXES)
    )
    if not parts:
        return []
    return [
        WorkbookDiagnostic(
            code="xlsx.unverified_ooxml_objects",
            severity="warning",
            message=(
                "工作簿包含当前文件级回归矩阵未验证的 OOXML 对象，"
                "导出后应人工抽查对象是否仍完整。"
            ),
            action=(
                "导出前保存一份原文件备份；若对象对业务重要，"
                "请在 Excel/WPS 中打开导出文件进行抽查。"
            ),
            details={"parts": parts[:100], "part_count": len(parts)},
        )
    ]


def _workbook_diagnostics(wb: Any, names: list[str]) -> list[WorkbookDiagnostic]:
    diagnostics = _rich_text_diagnostics(wb)
    external_parts = [name for name in names if name.startswith("xl/externalLinks/")]
    if external_parts or getattr(wb, "_external_links", None):
        diagnostics.append(
            WorkbookDiagnostic(
                code="xlsx.external_links",
                severity="warning",
                message="工作簿包含外部链接，链接目标不会在本地翻译流程中验证。",
                action="确认导出文件不依赖失效的外部链接，并在 Excel/WPS 中抽查。",
                details={"parts": external_parts[:100]},
            )
        )

    protected_sheets = [ws.title for ws in wb.worksheets if ws.protection.sheet]
    workbook_protected = bool(
        getattr(wb.security, "lockStructure", False)
        or getattr(wb.security, "lockWindows", False)
        or getattr(wb.security, "workbookPassword", None)
    )
    if protected_sheets or workbook_protected:
        diagnostics.append(
            WorkbookDiagnostic(
                code="xlsx.protection",
                severity="warning",
                message="工作簿或工作表启用了保护，导出后的编辑行为需要人工确认。",
                action="确认当前用户有权编辑导出文件；必要时先在 Excel/WPS 中解除保护。",
                details={
                    "workbook_protected": workbook_protected,
                    "protected_sheets": protected_sheets,
                },
            )
        )
    diagnostics.extend(_object_part_diagnostics(names))
    return diagnostics


def preflight_workbook(xlsx_path: str | os.PathLike[str]) -> WorkbookPreflight:
    """Validate extension, OOXML container and openpyxl readability.

    ``ok`` only describes whether task creation may proceed.  Warnings are
    retained in ``diagnostics`` so the caller can show them before export.
    Unsupported extensions and corrupt OOXML are distinct ``kind`` values:
    ``unsupported``, ``corrupt``, and ``missing``.  A valid workbook is
    returned as ``kind == 'valid'`` even when it has non-blocking warnings.
    """
    path = Path(xlsx_path)
    path_text = os.fspath(path)
    if not path.is_file():
        diagnostic = WorkbookDiagnostic(
            code="xlsx.input_missing",
            severity="error",
            message=f"输入文件不存在或不是普通文件: {path_text}",
            action="重新选择一个存在的 .xlsx 文件。",
        )
        return WorkbookPreflight(path_text, "missing", False, (diagnostic,))
    if path.suffix.lower() != ".xlsx":
        diagnostic = WorkbookDiagnostic(
            code="xlsx.unsupported_extension",
            severity="error",
            message=f"不支持的文件类型: {path.suffix or '无扩展名'}，当前只接受 .xlsx。",
            action="请在 Excel/WPS 中使用“另存为”转换为 .xlsx 后重试。",
            details={"extension": path.suffix},
        )
        return WorkbookPreflight(path_text, "unsupported", False, (diagnostic,))
    if not zipfile.is_zipfile(path):
        diagnostic = WorkbookDiagnostic(
            code="xlsx.invalid_zip",
            severity="error",
            message="文件扩展名是 .xlsx，但内容不是有效的 OOXML ZIP 文件。",
            action="重新导出工作簿；不要只修改文件扩展名。",
        )
        return WorkbookPreflight(path_text, "corrupt", False, (diagnostic,))

    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
            if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                raise ValueError("缺少 [Content_Types].xml 或 xl/workbook.xml")
        wb = openpyxl.load_workbook(
            path,
            data_only=False,
            keep_links=True,
            keep_vba=False,
            rich_text=True,
        )
    except Exception as exc:
        diagnostic = WorkbookDiagnostic(
            code="xlsx.corrupt_workbook",
            severity="error",
            message=f"无法打开工作簿，文件可能损坏或不是有效的 .xlsx: {exc}",
            action="用 Excel/WPS 打开并另存为新的 .xlsx 文件后重试。",
            details={"exception_type": type(exc).__name__},
        )
        return WorkbookPreflight(path_text, "corrupt", False, (diagnostic,))

    try:
        diagnostics = tuple(_workbook_diagnostics(wb, names))
    finally:
        wb.close()
    return WorkbookPreflight(path_text, "valid", True, diagnostics)


def require_valid_workbook(xlsx_path: str | os.PathLike[str]) -> WorkbookPreflight:
    """Return preflight data or raise an error carrying structured details."""
    result = preflight_workbook(xlsx_path)
    if not result.ok:
        raise WorkbookPreflightError(result)
    return result


def extract_artifact(xlsx_path: str | os.PathLike[str]) -> SourceArtifact:
    """Read one validated workbook and return its source artifact."""
    path = Path(xlsx_path)
    require_valid_workbook(path)
    input_sha256 = sha256_file(path)
    wb = openpyxl.load_workbook(
        path,
        data_only=False,
        keep_links=True,
        keep_vba=False,
        rich_text=True,
    )

    groups: list[_TextGroup] = []
    index_by_text: dict[str, int] = {}
    sheets_seen: list[str] = []
    cells_total = 0
    cells_translatable = 0
    try:
        for ws in wb.worksheets:
            sheets_seen.append(ws.title)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    cells_total += 1
                    if not _is_translatable(cell):
                        continue
                    cells_translatable += 1
                    text = cell_text(cell.value)
                    if text not in index_by_text:
                        group = _TextGroup(text=text, id=len(groups))
                        index_by_text[text] = group.id
                        groups.append(group)
                    else:
                        group = groups[index_by_text[text]]
                    group.cells.append(CellRef(ws.title, cell.coordinate))
    finally:
        wb.close()

    if sha256_file(path) != input_sha256:
        raise ArtifactError("提取过程中输入工作簿发生变化，请重试")

    stats = {
        "sheets": sheets_seen,
        "cells_total": cells_total,
        "cells_translatable": cells_translatable,
        "unique_texts": len(groups),
    }
    return SourceArtifact.create(
        input_sha256=input_sha256,
        items=(
            SourceItem(id=group.id, text=group.text, cells=tuple(group.cells))
            for group in groups
        ),
        stats=stats,
    )


def extract(
    xlsx_path: str | os.PathLike[str],
    txt_path: str | os.PathLike[str],
    json_path: str | os.PathLike[str],
) -> dict[str, Any]:
    """Internal format adapter for the GUI's structured extraction service."""
    artifact = extract_artifact(xlsx_path)
    source_text = "".join(f"{escape_text(item.text)}\n" for item in artifact.items)
    mapping = [
        {
            "id": item.id,
            "text": item.text,
            "cells": [cell.to_dict() for cell in item.cells],
        }
        for item in artifact.items
    ]
    atomic_write_text(txt_path, source_text)
    atomic_write_json(json_path, mapping)
    return {
        **artifact.stats,
        "source_revision": artifact.source_revision,
        "txt_path": os.fspath(txt_path),
        "json_path": os.fspath(json_path),
    }
